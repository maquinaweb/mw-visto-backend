from datetime import datetime
from io import BytesIO
from uuid import uuid4

from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class SheetService:
    """
    Serviço independente para geração de relatórios Excel
    Foca apenas na estruturação e geração do arquivo, recebendo dados prontos
    """

    def __init__(self):
        """Inicializa o serviço de relatórios"""
        pass

    def gerar_relatorio(
        self,
        titulo_planilha: str,
        headers: list,
        dados: list,
        formatters: list = None,
        planilhas_adicionais: list = None,
        nome_arquivo: str = None,
    ) -> dict:
        wb = Workbook()
        wb.iso_dates = True
        ws = wb.active

        ws.title = titulo_planilha
        self._escrever_cabecalhos(headers, ws)
        self._escrever_dados(dados, ws, formatters)

        if planilhas_adicionais:
            for planilha in planilhas_adicionais:
                ws_extra = wb.create_sheet(planilha["titulo"])
                self._escrever_cabecalhos(planilha["headers"], ws_extra)
                self._escrever_dados(planilha["dados"], ws_extra)

        ws.freeze_panes = "A2"

        self._ajustar_largura_colunas(ws)

        content_io = BytesIO()
        wb.save(content_io)
        content_io.seek(0)

        if not nome_arquivo:
            nome_arquivo = f"sheet_{uuid4()}"

        return ContentFile(content_io.getvalue(), name=f"{nome_arquivo}.xlsx")

    def _escrever_cabecalhos(self, headers: list, ws):
        alignment_header = Alignment(horizontal="center", vertical="center")
        font_header = Font(bold=True, color="FFFFFF")

        for index, header in enumerate(headers):
            cell = ws.cell(row=1, column=index + 1, value=header)
            cell.font = font_header
            cell.alignment = alignment_header
            cell.fill = PatternFill(
                start_color="333333", end_color="333333", fill_type="solid"
            )

    def _ajustar_largura_colunas(self, ws):
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    def _escrever_dados(self, dados: list, ws, formatters: list = None):
        for linha_index, linha_dados in enumerate(dados):
            for coluna_index, valor in enumerate(linha_dados):
                if isinstance(valor, datetime) and valor.tzinfo is not None:
                    valor = valor.replace(tzinfo=None)

                cell = ws.cell(
                    row=linha_index + 2, column=coluna_index + 1, value=valor
                )

                if formatters and coluna_index < len(formatters):
                    formatter_type = formatters[coluna_index]
                    if formatter_type == "currency":
                        cell.number_format = "R$ #,##0.00"
                    elif formatter_type == "number":
                        cell.number_format = "#,##0.00"
